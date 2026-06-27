import csv, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

src = sys.argv[1] if len(sys.argv) > 1 else "submission.csv"
dst = sys.argv[2] if len(sys.argv) > 2 else "submission.xlsx"

with open(src, encoding="utf-8") as f:
    rows = list(csv.reader(f))

wb = Workbook(); ws = wb.active; ws.title = "ranking"
for r in rows:
    ws.append(r)

# header styling
for c in ws[1]:
    c.font = Font(bold=True)
    c.alignment = Alignment(vertical="center")
ws.freeze_panes = "A2"

# numeric types for rank & score; column widths
for row in ws.iter_rows(min_row=2):
    row[1].value = int(row[1].value)       # rank
    row[2].value = float(row[2].value)     # score
    row[2].number_format = "0.0000"
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 7
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 120

wb.save(dst)
print(f"wrote {dst}: {len(rows)-1} candidates, columns {rows[0]}")